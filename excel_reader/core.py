from dataclasses import dataclass
from dataclasses_json import dataclass_json
from typing import Any, Dict, List
import csv
import os
from openpyxl import load_workbook

Row = Dict[str, Any]

@dataclass_json
@dataclass
class Sheet:
    name: str
    rows: List[Row]

@dataclass_json
@dataclass
class Workbook:
    name: str
    sheets: List[Sheet]

def read_excel(filepath: str) -> Workbook:
    wb = load_workbook(filepath, data_only=True)
    sheets: List[Sheet] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract values from rows
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            sheets.append(Sheet(name=sheet_name, rows=[]))
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        data_rows: List[Row] = []

        for row in rows[1:]:
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            data_rows.append(row_dict)

        sheets.append(Sheet(name=sheet_name, rows=data_rows))

    return Workbook(name=os.path.basename(filepath), sheets=sheets)

def read_csv(filepath: str) -> Sheet:
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows: List[Row] = [dict(row) for row in reader]

    sheet_name = os.path.splitext(os.path.basename(filepath))[0]
    return Sheet(name=sheet_name, rows=rows)

